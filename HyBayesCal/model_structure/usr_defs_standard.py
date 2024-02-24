"""
Instantiate global variables of user definitions made in user_input.xlsx (or any other file for another
software than Telemac) with this basic class foruser definitions required by Bayesian active learning

A new software implementation should contain a class UserDefsNewSoftware that inherits from this class.
See an example in usr_def_TELEMAC.UserDefsTelemac.
"""

import pandas as _pd
from openpyxl import load_workbook


class UserDefs:
    def __init__(self, input_worbook_name="user-input.xlsx", *args, **kwargs):
        self.input_xlsx_name = input_worbook_name  # never remove this, even if you do not use an XLSX file for user input defs
        print(
            "Using %s to read user settings. To change the user settings, run THIS.write_global_settings('path/to/workbook.xlsx')" % self.input_xlsx_name)

        self.file_write_dir = ""

        # initialize capital letter class variables that are defined through the user XLSX file
        self.CALIB_PAR_SET = {}  # dict for direct calibration optimization parameters
        self.CALIB_PTS = None  # numpy array to be loaded from calibration_points file
        self.CALIB_TARGETS = []  # list of calibration target (measurement) data (e.g. topographic change)
        self.init_runs = int()  # int number of initial full-complexity runs
        self.init_run_sampling = str()  # how the number of initial full-complexity runs should be sampled
        self.IT_LIMIT = int()  # int limit for Bayesian iterations
        self.MC_SAMPLES = int()  # int for Monte Carlo samples
        self.MC_SAMPLES_AL = int()  # int for Monte Carlo samples for active learning
        self.AL_SAMPLES = int()  # int for no. of active learning sampling size
        self.AL_STRATEGY = str()  # str for active learning strategy
        self.score_method = str()  # str for score calculation method to use in BAL_core.BAL.compute_bayesian_scores
        self.SIM_DIR = str()  # relative path for simulations - all paths must not end with any "/" or "\\"
        self.BME = None
        self.RE = None
        self.al_BME = None
        self.al_RE = None

    def assign_global_settings(self, *args, **kwargs):
        """Replace with method to assign Bayesian Calibration settings. Specifically, define
        the following variables:

        self.CALIB_PTS := calibration parameter points (str/file)
        self.AL_STRATEGY := active learning strategy (str) (Entropy
        self.score_method := score calculation method, either Bayesian weighting or rejection sampling (str)
        self.init_runs := number of initial full-complexity runs (int)
        self.init_run_sampling := strategy for determining calibration parameter values of initial runs (e.g., min-max)
        self.IT_LIMIT := maximum number of active learning iterations (int), typically 50-100
        self.AL_SAMPLES := no. of samples for moving from prior to posterior
        self.MC_SAMPLES := no. of uniform samples from calibration parameter ranges
        self.MC_SAMPLES_AL := number of monte carlo samples for exploring the output space

        Make sure to also assign all other variables defined in the __init__ method
        """
        pass

    def check_user_input(self):
        """Not absolutely necessary but recommended: verify if global variables are correctly assigned"""
        pass

    def read_wb_range(self, read_range, sheet_name="MAIN"):
        """Read a certain range of a workbook only with openpyxl

        :param str read_range: letter-number read range in workbook (e.g. "A2:B4")
        :param str sheet_name: name of the sheet to read (default is MAIN from user-inpux.xlsx)
        :return pd.DataFrame: xlsx contents in the defined range
        """
        ws = load_workbook(filename=self.input_xlsx_name, read_only=True, data_only=True)[sheet_name]
        # Read the cell values into a list of lists
        data_rows = []
        for row in ws[read_range]:
            data_rows.append([cell.value for cell in row])
            print(data_rows)
        return _pd.DataFrame(data_rows)
